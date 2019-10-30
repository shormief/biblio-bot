from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from slackclient import SlackClient

SLACK_BOT_TOKEN = ""
SLACK_CHANNEL = "#biblio-bot"
slack_client = SlackClient(SLACK_BOT_TOKEN)

def scraping_books():
    base_url = 'https://www.queenslibrary.org/search/book?searchField='
    base2_url = '&category=book&searchFilter='

    driver = webdriver.Chrome()

    loc = ("Wishlist.xlsx")
    wb = openpyxl.load_workbook(loc)
    sheet = wb.get_sheet_by_name('Sheet1')

    # indices of the row/columns
    i = 2 # the books are listed from row 2
    notifiedCol = 4
    isbnCol = 2
    titleCol = 1

    ''' LOOPING ''' 
    for x in range(i, sheet.max_row):
        if sheet.cell(x, notifiedCol).value == "No": 
            # title of the book
            title = (sheet.cell(x, titleCol)).value 

            # searching for the book 
            isbn = int((sheet.cell(x, isbnCol)).value)
            print(isbn)
            isbn = str(isbn)
            url = base_url + isbn + base2_url
            
            # opening up Chrome     
            driver.get(url)
            time.sleep(5)
            
            '''  CHECKING TO SEE IF A BOOK IS AVAILABLE FOR REQUEST  '''  
            # hovering over the book block 
            try: 
                element_to_hover_over = driver.find_element_by_class_name("searchResultEachBlock")
                ActionChains(driver).move_to_element(element_to_hover_over).perform()
            except: 
                continue
            driver.implicitly_wait(5)

            # clicking on quickViewSearch
            try: 
                quickView = driver.find_element_by_class_name("quickviewSearch")
                driver.execute_script("arguments[0].click();", quickView)
            except: 
                continue
            driver.implicitly_wait(20)

            # checking if the book is available for request
            try:
                objectFrame = driver.find_element_by_id("widgetObj")
                # checking if the book has available copies
                openWidget = objectFrame.get_attribute("data")
                time.sleep(10)
                driver.get(openWidget)
                ableToRequest = driver.find_element_by_id("requestLocationLogin")

                # getting the direct URL to request 
                requestURL = ableToRequest.get_attribute("href")

                # writing the message
                message = "*{}* is available for request \n".format(title)
                message += "{}".format(requestURL)

                # posting the message to slack
                slack_client.api_call (
                    "chat.postMessage", 
                    channel=SLACK_CHANNEL, 
                    text=message, 
                    username='Bibli'
                )
                
                # rewriting excel sheet 
                sheet.cell(x, notifiedCol).value = 'Yes'
                wb.save(loc)
            except: 
                continue
            time.sleep(20) 
    driver.close()
    return 
